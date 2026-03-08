"""Tests for create_xlsx tool."""

from __future__ import annotations

from pathlib import Path

import openpyxl

from office_agent.tools.xlsx_tool import create_xlsx


def test_create_xlsx_basic(sample_xlsx_plan: dict, tmp_path: Path) -> None:
    result = create_xlsx(plan=sample_xlsx_plan, out_dir=str(tmp_path))
    assert result["tool_used"] == "create_xlsx"
    out = Path(result["output_path"])
    assert out.exists()
    assert out.suffix == ".xlsx"


def test_create_xlsx_sheet_names(sample_xlsx_plan: dict, tmp_path: Path) -> None:
    result = create_xlsx(plan=sample_xlsx_plan, out_dir=str(tmp_path))
    wb = openpyxl.load_workbook(result["output_path"])
    expected = [s["name"] for s in sample_xlsx_plan["sheets"]]
    assert wb.sheetnames == expected


def test_create_xlsx_headers(sample_xlsx_plan: dict, tmp_path: Path) -> None:
    result = create_xlsx(plan=sample_xlsx_plan, out_dir=str(tmp_path))
    wb = openpyxl.load_workbook(result["output_path"])
    ws = wb.active or wb.worksheets[0]
    first_sheet = sample_xlsx_plan["sheets"][0]
    header_values = [ws.cell(1, c + 1).value for c in range(len(first_sheet["headers"]))]
    assert header_values == first_sheet["headers"]


def test_create_xlsx_formula(sample_xlsx_plan: dict, tmp_path: Path) -> None:
    result = create_xlsx(plan=sample_xlsx_plan, out_dir=str(tmp_path))
    wb = openpyxl.load_workbook(result["output_path"], data_only=False)
    ws = wb.worksheets[0]
    formula_spec = sample_xlsx_plan["sheets"][0]["formulas"][0]
    cell_val = ws[formula_spec["cell"]].value
    assert cell_val == formula_spec["formula"]


def test_tool_registry_xlsx(registry, sample_xlsx_plan: dict, tmp_path: Path) -> None:
    result = registry.call("create_xlsx", plan=sample_xlsx_plan, out_dir=str(tmp_path))
    assert result["tool_used"] == "create_xlsx"
