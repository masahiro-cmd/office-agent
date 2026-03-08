"""create_xlsx: Generate an Excel workbook from a plan JSON."""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def create_xlsx(plan: dict, out_dir: str, template_dir: str = "./templates") -> dict:
    """
    Generate a .xlsx workbook from a validated plan dict.

    Args:
        plan: Validated workbook plan (must conform to xlsx_schema.json).
        out_dir: Directory where the output file will be saved.
        template_dir: Not used for xlsx (kept for consistent signature).

    Returns:
        {"output_path": str, "tool_used": "create_xlsx"}
    """
    title: str = plan.get("workbook_title", "Untitled")
    sheets: list[dict] = plan.get("sheets", [])

    wb = openpyxl.Workbook()
    # Remove default sheet
    if wb.active and not sheets:
        pass  # keep it
    elif wb.active:
        wb.remove(wb.active)

    for sheet_plan in sheets:
        _add_sheet(wb, sheet_plan)

    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)
    safe_title = _safe_filename(title)
    file_path = out_path / f"{safe_title}.xlsx"
    wb.save(str(file_path))
    logger.info(f"Saved xlsx: {file_path}")

    return {"output_path": str(file_path), "tool_used": "create_xlsx"}


def _add_sheet(wb: openpyxl.Workbook, sheet_plan: dict) -> None:
    name = sheet_plan.get("name", "Sheet")
    headers: list[str] = sheet_plan.get("headers", [])
    rows: list[list] = sheet_plan.get("rows", [])
    formulas: list[dict] = sheet_plan.get("formulas") or []
    filters: bool = sheet_plan.get("filters", False)
    freeze_pane: str | None = sheet_plan.get("freeze_pane")

    ws = wb.create_sheet(title=name)

    # Headers
    if headers:
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT

    # Data rows
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Formulas
    for formula_spec in formulas:
        cell_ref: str = formula_spec.get("cell", "")
        formula: str = formula_spec.get("formula", "")
        if cell_ref and formula:
            ws[cell_ref] = formula

    # Auto-filter
    if filters and headers:
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last_col}1"

    # Freeze pane
    if freeze_pane:
        ws.freeze_panes = freeze_pane

    # Auto-width (approximate)
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def _safe_filename(name: str) -> str:
    invalid = r'\/:*?"<>|'
    for ch in invalid:
        name = name.replace(ch, "_")
    return name[:100]
